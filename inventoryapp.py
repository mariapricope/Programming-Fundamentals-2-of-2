# Main.py

from AddInv import add_item
from EditInv import edit_item
from RemInv import remove_item
from ViewInv import view_inventory

def display_menu():
    """Displays the main menu for the Inventory Management System."""
    print("\nInventory Management System")
    print("1. Add Item")
    print("2. Edit Item")
    print("3. View Inventory")
    print("4. Remove Item")
    print("5. Exit")

def main():
    """Main function that runs the inventory system in a loop until the user exits."""
    while True:
        display_menu()
        choice = input("Enter your choice (1-5) or type 'exit' to abort: ").strip().lower()
        if choice == '1':
            add_item()
        elif choice == '2':
            edit_item()
        elif choice == '3':
            view_inventory()
        elif choice == '4':
            remove_item()
        elif choice == '5' or choice == 'exit':
            print("Exiting the system.")
            break  # Exit the loop and end the program
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

if __name__ == "__main__":
    main()  # Run the main function if this file is executed



# AddInv.py
def add_item():
    """Adds a new item to the inventory."""
    from Inventory import import_inventory, export_inventory

    name = input("Enter item name: ").strip()
    quantity = int(input("Enter item quantity: ").strip())
    price = float(input("Enter item price: ").strip())

    inventory = import_inventory()

    # Check if item already exists
    existing_item = next((item for item in inventory if item[1].strip().lower() == name.lower()), None)

    if existing_item:
        item_id, old_name, old_quantity, old_price = existing_item
        # Update the existing item
        inventory = [(item_id, name, old_quantity + quantity, price) if item[0] == item_id else item for item in inventory]
    else:
        # Assign new ID
        next_id = max(item[0] for item in inventory) + 1 if inventory else 1
        inventory.append((next_id, name, quantity, price))

    export_inventory(inventory)
    print(f"Item '{name}' added/updated successfully.")



# EditInv.py
def edit_item():
    """Edits an existing item in the inventory by name."""
    from Inventory import import_inventory, export_inventory

    inventory = import_inventory()

    if not inventory:
        print("Inventory is empty. No item to edit.")
        return

    name = input("Enter item name to edit (or type 'exit' to cancel): ").strip().lower()
    if name == "exit":
        print("Edit operation cancelled.")
        return

    # Find item by name
    item_found = False
    for i, item in enumerate(inventory):
        if item[1].strip().lower() == name:
            item_found = True
            print(f"Current details - Name: {item[1]}, Quantity: {item[2]}, Price: £{item[3]:.2f}")
            new_name = input("Enter new name (leave blank to keep current): ").strip()
            new_quantity = input("Enter new quantity (leave blank to keep current): ").strip()
            new_price = input("Enter new price (leave blank to keep current): ").strip()

            # Update item details if provided
            if new_name:
                inventory[i] = (item[0], new_name, inventory[i][2], inventory[i][3])
            if new_quantity:
                inventory[i] = (item[0], inventory[i][1], int(new_quantity), inventory[i][3])
            if new_price:
                inventory[i] = (item[0], inventory[i][1], inventory[i][2], float(new_price))

            export_inventory(inventory)
            print(f"Item ID {item[0]} updated successfully.")
            break

    if not item_found:
        print(f"No items with the name '{name}' found.")


# ViewInv.py
def edit_item():
    """Edits an existing item in the inventory by name."""
    from Inventory import import_inventory, export_inventory

    inventory = import_inventory()

    if not inventory:
        print("Inventory is empty. No item to edit.")
        return

    name = input("Enter item name to edit (or type 'exit' to cancel): ").strip().lower()
    if name == "exit":
        print("Edit operation cancelled.")
        return

    # Find item by name
    item_found = False
    for i, item in enumerate(inventory):
        if item[1].strip().lower() == name:
            item_found = True
            print(f"Current details - Name: {item[1]}, Quantity: {item[2]}, Price: £{item[3]:.2f}")
            new_name = input("Enter new name (leave blank to keep current): ").strip()
            new_quantity = input("Enter new quantity (leave blank to keep current): ").strip()
            new_price = input("Enter new price (leave blank to keep current): ").strip()

            # Update item details if provided
            if new_name:
                inventory[i] = (item[0], new_name, inventory[i][2], inventory[i][3])
            if new_quantity:
                inventory[i] = (item[0], inventory[i][1], int(new_quantity), inventory[i][3])
            if new_price:
                inventory[i] = (item[0], inventory[i][1], inventory[i][2], float(new_price))

            export_inventory(inventory)
            print(f"Item ID {item[0]} updated successfully.")
            break

    if not item_found:
        print(f"No items with the name '{name}' found.")




# RemInv.py
def remove_item():
    """Removes all records of an item from the inventory based on item name."""
    from Inventory import import_inventory, export_inventory

    inventory = import_inventory()

    if not inventory:
        print("Inventory is empty. No item to remove.")
        return

    name = input("Enter item name to remove (or type 'exit' to cancel): ").strip().lower()
    if name == "exit":
        print("Remove operation cancelled.")
        return

    # Create a new inventory list excluding all items with the specified name
    new_inventory = [item for item in inventory if item[1].strip().lower() != name]

    if len(new_inventory) == len(inventory):
        print(f"No items with the name '{name}' were found in the inventory.")
    else:
        print(f"All items with the name '{name}' have been removed.")
        export_inventory(new_inventory)


# Inventory.py
import openpyxl

def import_inventory(excel_file='inventory.xlsx'):
    """Imports inventory from an Excel file."""
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        inventory = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            inventory.append(row)
        return inventory
    except FileNotFoundError:
        print("No existing Excel file found. Starting with an empty inventory.")
        return []

def export_inventory(inventory, excel_file='inventory.xlsx'):
    """Exports inventory to an Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"
    sheet.append(["Item ID", "Name", "Quantity", "Price"])
    for item in inventory:
        sheet.append(item)
    workbook.save(excel_file)
    print("Inventory saved to Excel file.")





from AddInv import add_item
from EditInv import edit_item
from RemInv import remove_item
from ViewInv import view_inventory

def display_menu():
    """Displays the main menu for the Inventory Management System."""
    print("\nInventory Management System")
    print("1. Add Item")
    print("2. Edit Item")
    print("3. View Inventory")
    print("4. Remove Item")
    print("5. Exit")

def main():
    """Main function that runs the inventory system in a loop until the user exits."""
    while True:
        display_menu()
        choice = input("Enter your choice (1-5) or type 'exit' to abort: ").strip().lower()
        if choice == '1':
            add_item()
        elif choice == '2':
            edit_item()
        elif choice == '3':
            view_inventory()
        elif choice == '4':
            remove_item()
        elif choice == '5' or choice == 'exit':
            print("Exiting the system.")
            break  # Exit the loop and end the program
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

if __name__ == "__main__":
    main()  # Run the main function if this file is executed
