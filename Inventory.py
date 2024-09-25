# Inventory.py
import openpyxl


class Inventory:
    """
    Inventory class for handling Excel-based inventory storage and management.
    """

    def __init__(self, excel_file='inventory.xlsx'):
        """
        Initializes the Inventory class with the Excel file.
        
        :param excel_file: The name of the Excel file to store the inventory (default is 'inventory.xlsx').
        """
        self.excel_file = excel_file
        self.inventory = self.import_inventory()

    def import_inventory(self):
        """
        Imports the inventory from the Excel file.
        
        :return: A list of inventory items.
        """
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active
            inventory = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
            return inventory
        except FileNotFoundError:
            print("No existing Excel file found. Starting with an empty inventory.")
            return []

    def export_inventory(self):
        """
        Exports the current inventory to the Excel file, overwriting the file.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        sheet.append(["Item ID", "Name", "Quantity", "Price"])

        for item in self.inventory:
            sheet.append(item)

        workbook.save(self.excel_file)
        print("Inventory saved to Excel file.")
